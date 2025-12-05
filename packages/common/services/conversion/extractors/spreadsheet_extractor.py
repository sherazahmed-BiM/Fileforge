"""
Spreadsheet Extractor for FileForge

Extracts data from Excel (XLSX/XLS) and CSV files using openpyxl and pandas.
"""

import csv
from pathlib import Path
from typing import Any, Optional

try:
    import pandas as pd
except ImportError:
    pd = None

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

from packages.common.services.conversion.extractors.base import (
    BaseExtractor,
    ElementType,
    ExtractedElement,
    ExtractionResult,
)
from packages.common.core.logging import get_logger


logger = get_logger(__name__)


class XLSXExtractor(BaseExtractor):
    """
    Extractor for Excel files (XLSX) using openpyxl.

    Features:
    - Sheet-by-sheet extraction
    - Table structure preservation
    - Sheet metadata
    """

    SUPPORTED_EXTENSIONS = {".xlsx", ".xls"}

    def __init__(self):
        """Initialize XLSX extractor."""
        super().__init__()

        if load_workbook is None:
            raise ImportError(
                "openpyxl is required for XLSX extraction. "
                "Install with: pip install openpyxl"
            )

    def extract(
        self,
        file_path: str | Path,
        max_rows: Optional[int] = None,
        **options: Any,
    ) -> ExtractionResult:
        """
        Extract content from an Excel file.

        Args:
            file_path: Path to the Excel file
            max_rows: Maximum rows to extract per sheet (None = all)

        Returns:
            ExtractionResult with all extracted elements
        """
        file_path = Path(file_path)
        logger.info(f"Extracting XLSX: {file_path.name}")

        elements: list[ExtractedElement] = []
        warnings: list[str] = []
        metadata: dict[str, Any] = {}

        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
        except Exception as e:
            logger.error(f"Failed to open XLSX: {e}")
            raise

        try:
            sheet_names = wb.sheetnames
            metadata["sheet_count"] = len(sheet_names)
            metadata["sheet_names"] = sheet_names

            for sheet_name in sheet_names:
                ws = wb[sheet_name]
                sheet_element = self._extract_sheet(ws, sheet_name, max_rows)
                if sheet_element:
                    elements.append(sheet_element)

        finally:
            wb.close()

        # Compute raw text
        raw_text = "\n\n".join(
            el.content for el in elements
        )

        word_count = self._count_words(raw_text)

        return ExtractionResult(
            elements=elements,
            metadata=metadata,
            raw_text=raw_text,
            page_count=len(sheet_names),
            word_count=word_count,
            extraction_method="openpyxl",
            warnings=warnings,
        )

    def _extract_sheet(
        self,
        ws: Any,
        sheet_name: str,
        max_rows: Optional[int] = None,
    ) -> Optional[ExtractedElement]:
        """Extract data from a single worksheet."""
        table_data: list[list[str]] = []
        row_count = 0

        for row in ws.iter_rows(values_only=True):
            if max_rows and row_count >= max_rows:
                break

            # Convert all cells to strings
            row_data = [str(cell) if cell is not None else "" for cell in row]

            # Skip completely empty rows
            if any(cell for cell in row_data):
                table_data.append(row_data)

            row_count += 1

        if not table_data:
            return None

        # Convert to text representation
        table_text = f"## Sheet: {sheet_name}\n\n" + self._table_to_markdown(table_data)

        return ExtractedElement(
            element_type=ElementType.TABLE,
            content=table_text,
            section=sheet_name,
            table_data=table_data,
            table_html=self._table_to_html(table_data),
            metadata={
                "sheet_name": sheet_name,
                "row_count": len(table_data),
                "column_count": max(len(row) for row in table_data) if table_data else 0,
            },
        )

    def _table_to_markdown(self, table_data: list[list[str]]) -> str:
        """Convert table data to markdown format."""
        if not table_data:
            return ""

        # Normalize column count
        max_cols = max(len(row) for row in table_data)
        normalized = [row + [""] * (max_cols - len(row)) for row in table_data]

        lines = []

        # Header row
        header = " | ".join(str(cell) for cell in normalized[0])
        lines.append(f"| {header} |")
        lines.append("| " + " | ".join("---" for _ in range(max_cols)) + " |")

        # Data rows
        for row in normalized[1:]:
            row_text = " | ".join(str(cell) for cell in row)
            lines.append(f"| {row_text} |")

        return "\n".join(lines)

    def _table_to_html(self, table_data: list[list[str]]) -> str:
        """Convert table data to HTML format."""
        if not table_data:
            return ""

        # Normalize column count
        max_cols = max(len(row) for row in table_data)
        normalized = [row + [""] * (max_cols - len(row)) for row in table_data]

        html_parts = ["<table>"]

        # Header
        html_parts.append("<thead><tr>")
        for cell in normalized[0]:
            html_parts.append(f"<th>{cell}</th>")
        html_parts.append("</tr></thead>")

        # Body
        html_parts.append("<tbody>")
        for row in normalized[1:]:
            html_parts.append("<tr>")
            for cell in row:
                html_parts.append(f"<td>{cell}</td>")
            html_parts.append("</tr>")
        html_parts.append("</tbody></table>")

        return "".join(html_parts)


class CSVExtractor(BaseExtractor):
    """
    Extractor for CSV files using Python's csv module and pandas.

    Features:
    - Automatic delimiter detection
    - Encoding handling
    - Table structure preservation
    """

    SUPPORTED_EXTENSIONS = {".csv", ".tsv"}

    def __init__(self):
        """Initialize CSV extractor."""
        super().__init__()

    def extract(
        self,
        file_path: str | Path,
        max_rows: Optional[int] = None,
        encoding: Optional[str] = None,
        delimiter: Optional[str] = None,
        **options: Any,
    ) -> ExtractionResult:
        """
        Extract content from a CSV file.

        Args:
            file_path: Path to the CSV file
            max_rows: Maximum rows to extract (None = all)
            encoding: File encoding (auto-detected if None)
            delimiter: CSV delimiter (auto-detected if None)

        Returns:
            ExtractionResult with all extracted elements
        """
        file_path = Path(file_path)
        logger.info(f"Extracting CSV: {file_path.name}")

        elements: list[ExtractedElement] = []
        warnings: list[str] = []
        metadata: dict[str, Any] = {}

        # Try to detect encoding
        if encoding is None:
            encoding = self._detect_encoding(file_path)

        # Try to detect delimiter
        if delimiter is None:
            delimiter = self._detect_delimiter(file_path, encoding)

        try:
            table_data = self._read_csv(
                file_path, encoding, delimiter, max_rows
            )
        except Exception as e:
            logger.error(f"Failed to read CSV: {e}")
            raise

        if table_data:
            metadata["row_count"] = len(table_data)
            metadata["column_count"] = max(len(row) for row in table_data) if table_data else 0
            metadata["encoding"] = encoding
            metadata["delimiter"] = delimiter

            table_text = self._table_to_markdown(table_data)

            elements.append(
                ExtractedElement(
                    element_type=ElementType.TABLE,
                    content=table_text,
                    table_data=table_data,
                    table_html=self._table_to_html(table_data),
                    metadata=metadata.copy(),
                )
            )

        raw_text = table_text if table_data else ""
        word_count = self._count_words(raw_text)

        return ExtractionResult(
            elements=elements,
            metadata=metadata,
            raw_text=raw_text,
            page_count=1,
            word_count=word_count,
            extraction_method="csv",
            warnings=warnings,
        )

    def _detect_encoding(self, file_path: Path) -> str:
        """Detect file encoding."""
        encodings = ["utf-8", "utf-8-sig", "latin-1", "cp1252"]

        for enc in encodings:
            try:
                with open(file_path, "r", encoding=enc) as f:
                    f.read(1024)
                return enc
            except UnicodeDecodeError:
                continue

        return "utf-8"  # Default fallback

    def _detect_delimiter(self, file_path: Path, encoding: str) -> str:
        """Detect CSV delimiter."""
        try:
            with open(file_path, "r", encoding=encoding) as f:
                sample = f.read(4096)
                sniffer = csv.Sniffer()
                dialect = sniffer.sniff(sample)
                return dialect.delimiter
        except Exception:
            # Default to comma for .csv, tab for .tsv
            if file_path.suffix.lower() == ".tsv":
                return "\t"
            return ","

    def _read_csv(
        self,
        file_path: Path,
        encoding: str,
        delimiter: str,
        max_rows: Optional[int] = None,
    ) -> list[list[str]]:
        """Read CSV file into list of rows."""
        table_data: list[list[str]] = []

        with open(file_path, "r", encoding=encoding, newline="") as f:
            reader = csv.reader(f, delimiter=delimiter)
            for i, row in enumerate(reader):
                if max_rows and i >= max_rows:
                    break
                # Convert all cells to strings and strip whitespace
                row_data = [str(cell).strip() for cell in row]
                table_data.append(row_data)

        return table_data

    def _table_to_markdown(self, table_data: list[list[str]]) -> str:
        """Convert table data to markdown format."""
        if not table_data:
            return ""

        # Normalize column count
        max_cols = max(len(row) for row in table_data) if table_data else 0
        normalized = [row + [""] * (max_cols - len(row)) for row in table_data]

        lines = []

        # Header row
        if normalized:
            header = " | ".join(str(cell) for cell in normalized[0])
            lines.append(f"| {header} |")
            lines.append("| " + " | ".join("---" for _ in range(max_cols)) + " |")

            # Data rows
            for row in normalized[1:]:
                row_text = " | ".join(str(cell) for cell in row)
                lines.append(f"| {row_text} |")

        return "\n".join(lines)

    def _table_to_html(self, table_data: list[list[str]]) -> str:
        """Convert table data to HTML format."""
        if not table_data:
            return ""

        # Normalize column count
        max_cols = max(len(row) for row in table_data) if table_data else 0
        normalized = [row + [""] * (max_cols - len(row)) for row in table_data]

        html_parts = ["<table>"]

        if normalized:
            # Header
            html_parts.append("<thead><tr>")
            for cell in normalized[0]:
                html_parts.append(f"<th>{cell}</th>")
            html_parts.append("</tr></thead>")

            # Body
            html_parts.append("<tbody>")
            for row in normalized[1:]:
                html_parts.append("<tr>")
                for cell in row:
                    html_parts.append(f"<td>{cell}</td>")
                html_parts.append("</tr>")
            html_parts.append("</tbody></table>")

        return "".join(html_parts)
