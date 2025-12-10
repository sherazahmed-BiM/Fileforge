"""
Docling-based Extractor for FileForge

Uses Docling for high-accuracy document text and image extraction.
Supports multiple file formats with format-specific optimizations.

Supported formats and extraction strategies:
- PDF: Full layout analysis, tables, images, OCR support
- DOCX: Word documents with formatting preserved
- XLSX: Excel spreadsheets as structured JSON with schema detection
- PPTX: PowerPoint presentations with slide structure
- HTML/HTM/XHTML: Web pages with structure preserved
- MD/Markdown: Markdown files with code block support
- CSV: Tabular data converted to JSON array of objects
- Images (PNG, JPG, JPEG, TIFF, BMP, WEBP): OCR text extraction
- AsciiDoc: Documentation format
"""

import base64
import csv
import re
from pathlib import Path
from typing import Any, Optional

from packages.common.services.conversion.extractors.base import (
    BaseExtractor,
    ElementType,
    ExtractedElement,
    ExtractionResult,
    StructuredTableData,
    TableSchema,
)
from packages.common.core.logging import get_logger

logger = get_logger(__name__)

# Lazy imports for Docling
_docling_available = None
_DocumentConverter = None
_PdfFormatOption = None
_WordFormatOption = None
_ExcelFormatOption = None
_PowerpointFormatOption = None
_HTMLFormatOption = None
_MarkdownFormatOption = None
_ImageFormatOption = None
_PdfPipelineOptions = None
_InputFormat = None
_ImageRefMode = None
_HTMLBackendOptions = None
_MarkdownBackendOptions = None


def _load_docling():
    """Lazy load Docling to avoid import overhead."""
    global _docling_available, _DocumentConverter, _PdfFormatOption
    global _WordFormatOption, _ExcelFormatOption, _PowerpointFormatOption
    global _HTMLFormatOption, _MarkdownFormatOption, _ImageFormatOption
    global _PdfPipelineOptions, _InputFormat, _ImageRefMode
    global _HTMLBackendOptions, _MarkdownBackendOptions

    if _docling_available is not None:
        return _docling_available

    try:
        from docling.document_converter import (
            DocumentConverter,
            PdfFormatOption,
            WordFormatOption,
            ExcelFormatOption,
            PowerpointFormatOption,
            HTMLFormatOption,
            MarkdownFormatOption,
            ImageFormatOption,
        )
        from docling.datamodel.pipeline_options import PdfPipelineOptions
        from docling.datamodel.base_models import InputFormat
        from docling_core.types.doc import ImageRefMode

        _DocumentConverter = DocumentConverter
        _PdfFormatOption = PdfFormatOption
        _WordFormatOption = WordFormatOption
        _ExcelFormatOption = ExcelFormatOption
        _PowerpointFormatOption = PowerpointFormatOption
        _HTMLFormatOption = HTMLFormatOption
        _MarkdownFormatOption = MarkdownFormatOption
        _ImageFormatOption = ImageFormatOption
        _PdfPipelineOptions = PdfPipelineOptions
        _InputFormat = InputFormat
        _ImageRefMode = ImageRefMode

        # Try to load backend options (may not be available in all versions)
        try:
            from docling.datamodel.backend_options import (
                HTMLBackendOptions,
                MarkdownBackendOptions,
            )
            _HTMLBackendOptions = HTMLBackendOptions
            _MarkdownBackendOptions = MarkdownBackendOptions
        except ImportError:
            _HTMLBackendOptions = None
            _MarkdownBackendOptions = None

        _docling_available = True
        logger.info("Docling loaded successfully with all format options")
    except ImportError as e:
        logger.warning(f"Docling not available: {e}")
        _docling_available = False

    return _docling_available


def _get_input_format(file_path: Path) -> Optional[Any]:
    """Map file extension to Docling InputFormat."""
    if _InputFormat is None:
        return None

    ext = file_path.suffix.lower()

    format_map = {
        # Documents
        '.pdf': _InputFormat.PDF,
        '.docx': _InputFormat.DOCX,
        '.xlsx': _InputFormat.XLSX,
        '.pptx': _InputFormat.PPTX,
        # Markup
        '.html': _InputFormat.HTML,
        '.htm': _InputFormat.HTML,
        '.xhtml': _InputFormat.HTML,
        '.md': _InputFormat.MD,
        '.markdown': _InputFormat.MD,
        '.adoc': _InputFormat.ASCIIDOC,
        '.asciidoc': _InputFormat.ASCIIDOC,
        # Data
        '.csv': _InputFormat.CSV,
        # Images (all use IMAGE type)
        '.png': _InputFormat.IMAGE,
        '.jpg': _InputFormat.IMAGE,
        '.jpeg': _InputFormat.IMAGE,
        '.tiff': _InputFormat.IMAGE,
        '.tif': _InputFormat.IMAGE,
        '.bmp': _InputFormat.IMAGE,
        '.webp': _InputFormat.IMAGE,
        '.gif': _InputFormat.IMAGE,
    }

    return format_map.get(ext)


class DoclingExtractor(BaseExtractor):
    """
    Multi-format extractor using Docling for high-accuracy extraction.

    Supports: PDF, DOCX, XLSX, PPTX, HTML, Markdown, CSV, images (with OCR).

    Each format is configured with optimal settings:
    - PDF: OCR, table structure, image extraction
    - DOCX: Word document parsing with structure
    - XLSX: Spreadsheet to structured JSON
    - PPTX: Slide-by-slide extraction
    - HTML: Web content with image parsing
    - Markdown: Code block preservation
    - CSV: Structured JSON with type inference
    - Images: OCR text extraction
    """

    SUPPORTED_EXTENSIONS = {
        # Documents
        ".pdf", ".docx", ".xlsx", ".pptx",
        # Markup
        ".html", ".htm", ".xhtml", ".md", ".markdown", ".adoc", ".asciidoc",
        # Data
        ".csv",
        # Images
        ".png", ".jpg", ".jpeg", ".tiff", ".tif", ".bmp", ".webp", ".gif",
    }

    def __init__(
        self,
        enable_ocr: bool | None = None,
        enable_tables: bool = True,
        generate_images: bool = True,
        images_scale: float = 2.0,
        enable_picture_description: bool | None = None,
        vlm_model: str | None = None,
    ):
        """
        Initialize Docling extractor for multiple file formats.

        Settings can be controlled via environment variables:
            DOCLING_ENABLE_OCR: "true"/"false" (default: true)
            DOCLING_ENABLE_VLM: "true"/"false" (default: false - requires GPU)
            DOCLING_VLM_MODEL: "granite", "smolvlm", or "none" (default: granite)

        Args:
            enable_ocr: Enable OCR for scanned documents and images (default: True)
            enable_tables: Enable table structure extraction
            generate_images: Extract images from documents
            images_scale: Scale factor for extracted images
            enable_picture_description: Use VLM to generate picture descriptions
            vlm_model: VLM model to use - "granite" or "smolvlm"
        """
        import os

        super().__init__()

        if not _load_docling():
            raise ImportError(
                "Docling is required for this extractor. "
                "Install with: pip install docling[easyocr]"
            )

        def env_bool(key: str, default: bool) -> bool:
            val = os.environ.get(key, "").lower()
            if val in ("true", "1", "yes"):
                return True
            elif val in ("false", "0", "no"):
                return False
            return default

        self.enable_ocr = (
            enable_ocr if enable_ocr is not None
            else env_bool("DOCLING_ENABLE_OCR", True)
        )
        self.enable_tables = enable_tables
        self.generate_images = generate_images
        self.images_scale = images_scale
        self.enable_picture_description = (
            enable_picture_description if enable_picture_description is not None
            else env_bool("DOCLING_ENABLE_VLM", False)
        )
        self.vlm_model = vlm_model or os.environ.get("DOCLING_VLM_MODEL", "granite")

        if self.enable_picture_description:
            logger.info(f"VLM picture description enabled with model: {self.vlm_model}")

        # Cache converters per format
        self._converters: dict[str, Any] = {}

    def _get_converter(self, file_path: Path):
        """Get or create converter with format-specific configuration."""
        input_format = _get_input_format(file_path)

        if input_format is None:
            raise ValueError(f"Unsupported file format: {file_path.suffix}")

        format_key = str(input_format)

        if format_key not in self._converters:
            format_options = {}

            # PDF: Full pipeline with OCR, tables, images
            if input_format == _InputFormat.PDF:
                pipeline_options = _PdfPipelineOptions()
                pipeline_options.do_ocr = self.enable_ocr
                pipeline_options.do_table_structure = self.enable_tables
                pipeline_options.generate_picture_images = self.generate_images
                pipeline_options.images_scale = self.images_scale

                if self.enable_picture_description and self.vlm_model != "none":
                    pipeline_options.do_picture_description = True
                    pipeline_options = self._configure_vlm_options(pipeline_options)

                format_options[input_format] = _PdfFormatOption(
                    pipeline_options=pipeline_options
                )

            # DOCX: Word document processing
            elif input_format == _InputFormat.DOCX:
                format_options[input_format] = _WordFormatOption()

            # XLSX: Excel spreadsheet processing
            elif input_format == _InputFormat.XLSX:
                format_options[input_format] = _ExcelFormatOption()

            # PPTX: PowerPoint processing
            elif input_format == _InputFormat.PPTX:
                format_options[input_format] = _PowerpointFormatOption()

            # HTML: Web content with image parsing
            elif input_format == _InputFormat.HTML:
                if _HTMLBackendOptions is not None:
                    backend_options = _HTMLBackendOptions(parse_images=True)
                    format_options[input_format] = _HTMLFormatOption(
                        backend_options=backend_options
                    )
                else:
                    format_options[input_format] = _HTMLFormatOption()

            # Markdown: Code block preservation
            elif input_format == _InputFormat.MD:
                if _MarkdownBackendOptions is not None:
                    backend_options = _MarkdownBackendOptions(keep_code_blocks=True)
                    format_options[input_format] = _MarkdownFormatOption(
                        backend_options=backend_options
                    )
                else:
                    format_options[input_format] = _MarkdownFormatOption()

            # Images: OCR pipeline
            elif input_format == _InputFormat.IMAGE:
                format_options[input_format] = _ImageFormatOption()

            # Create converter with format options
            if format_options:
                self._converters[format_key] = _DocumentConverter(
                    format_options=format_options
                )
            else:
                self._converters[format_key] = _DocumentConverter()

        return self._converters[format_key]

    def _configure_vlm_options(self, pipeline_options):
        """Configure VLM options for picture description."""
        try:
            if self.vlm_model == "granite":
                from docling.datamodel.pipeline_options import granite_picture_description
                pipeline_options.picture_description_options = granite_picture_description
                pipeline_options.picture_description_options.prompt = (
                    "Describe this image in detail. Include any text, diagrams, charts, "
                    "or visual elements. Be accurate and comprehensive."
                )
                logger.info("VLM enabled: Using Granite Vision for picture descriptions")

            elif self.vlm_model == "smolvlm":
                from docling.datamodel.pipeline_options import smolvlm_picture_description
                pipeline_options.picture_description_options = smolvlm_picture_description
                pipeline_options.picture_description_options.prompt = (
                    "Describe this image in detail. Include any text, diagrams, charts, "
                    "or visual elements. Be accurate and comprehensive."
                )
                logger.info("VLM enabled: Using SmolVLM for picture descriptions")

            else:
                logger.warning(
                    f"Unknown VLM model: {self.vlm_model}, disabling picture description"
                )
                pipeline_options.do_picture_description = False

        except ImportError as e:
            logger.warning(f"VLM support not available: {e}. Disabling picture description.")
            pipeline_options.do_picture_description = False
        except Exception as e:
            logger.warning(f"Failed to configure VLM: {e}. Disabling picture description.")
            pipeline_options.do_picture_description = False

        return pipeline_options

    def extract(
        self,
        file_path: str | Path,
        extract_images: bool = True,
        **options: Any,
    ) -> ExtractionResult:
        """
        Extract text and images from a file using Docling.

        Args:
            file_path: Path to the file
            extract_images: Whether to extract images

        Returns:
            ExtractionResult with extracted text and images
        """
        file_path = Path(file_path)
        file_ext = file_path.suffix.lower()

        input_format = _get_input_format(file_path)
        if input_format is None:
            raise ValueError(
                f"Unsupported file format: {file_ext}. "
                f"Supported formats: {', '.join(sorted(self.SUPPORTED_EXTENSIONS))}"
            )

        logger.info(f"Extracting {input_format.value.upper()} using Docling: {file_path.name}")

        # Route to format-specific handler
        if input_format == _InputFormat.CSV:
            return self._extract_csv(file_path)
        elif input_format == _InputFormat.XLSX:
            return self._extract_xlsx(file_path)
        elif input_format == _InputFormat.DOCX:
            return self._extract_docx(file_path, extract_images)
        elif input_format == _InputFormat.PPTX:
            return self._extract_pptx(file_path, extract_images)
        elif input_format == _InputFormat.HTML:
            return self._extract_html(file_path, extract_images)
        elif input_format == _InputFormat.MD:
            return self._extract_markdown(file_path)
        elif input_format == _InputFormat.IMAGE:
            return self._extract_image(file_path)
        elif input_format == _InputFormat.PDF:
            return self._extract_pdf(file_path, extract_images)
        elif input_format == _InputFormat.ASCIIDOC:
            return self._extract_asciidoc(file_path)
        else:
            return self._extract_generic(file_path, extract_images)

    def _extract_pdf(
        self,
        file_path: Path,
        extract_images: bool = True,
    ) -> ExtractionResult:
        """Extract content from PDF with full pipeline."""
        elements: list[ExtractedElement] = []
        warnings: list[str] = []
        metadata: dict[str, Any] = {}

        try:
            converter = self._get_converter(file_path)
            result = converter.convert(str(file_path))
            doc = result.document

            page_count = result.input.page_count if result.input else 0
            metadata = self._extract_metadata(result)
            metadata["format"] = "pdf"

            # Process all document items
            page_elements: dict[int, list[tuple[int, ExtractedElement]]] = {}
            order_counter = 0
            # Track seen image data to avoid duplicates (use first 200 chars as hash)
            seen_image_hashes: set[str] = set()

            for item, level in doc.iterate_items():
                item_page = self._get_item_page(item)
                item_order = self._get_item_order(item, order_counter)
                order_counter += 1

                if item_page not in page_elements:
                    page_elements[item_page] = []

                if self._is_picture_item(item):
                    if extract_images and self.generate_images:
                        img_element = self._extract_picture_item(item, order_counter, doc)
                        if img_element and img_element.image_data:
                            # Deduplicate based on image data hash
                            img_hash = img_element.image_data[:200]
                            if img_hash not in seen_image_hashes:
                                seen_image_hashes.add(img_hash)
                                # Add position to metadata for frontend ordering
                                img_element.metadata["position"] = item_order
                                page_elements[item_page].append((item_order, img_element))
                elif hasattr(item, 'text') and item.text:
                    text_el = ExtractedElement(
                        element_type=ElementType.TEXT,
                        content=item.text,
                        page_number=item_page,
                    )
                    page_elements[item_page].append((item_order, text_el))

            # Process pictures from doc.pictures only if we haven't found any images yet
            # This is a fallback for documents where iterate_items doesn't yield pictures
            if extract_images and self.generate_images and hasattr(doc, 'pictures'):
                for idx, picture in enumerate(doc.pictures):
                    picture_page = self._get_item_page(picture)
                    picture_order = self._get_item_order(picture, 1000 + idx)

                    if picture_page not in page_elements:
                        page_elements[picture_page] = []

                    img_element = self._extract_picture(picture, idx, doc)
                    if img_element and img_element.image_data:
                        # Deduplicate based on image data hash
                        img_hash = img_element.image_data[:200]
                        if img_hash not in seen_image_hashes:
                            seen_image_hashes.add(img_hash)
                            # Add position to metadata for frontend ordering
                            img_element.metadata["position"] = picture_order
                            page_elements[picture_page].append((picture_order, img_element))

            # Export markdown with embedded images
            try:
                markdown_text = self._export_markdown_with_images(doc)
                if markdown_text and markdown_text.strip():
                    metadata["markdown"] = markdown_text
            except Exception as e:
                logger.warning(f"Markdown export failed: {e}")

            # Flatten page elements
            for page_num in sorted(page_elements.keys()):
                sorted_elements = sorted(page_elements[page_num], key=lambda x: x[0])
                for _, el in sorted_elements:
                    elements.append(el)

            # Fallback to markdown if no elements
            if not elements:
                try:
                    markdown_text = self._export_markdown_with_images(doc)
                    if markdown_text and markdown_text.strip():
                        elements.append(
                            ExtractedElement(
                                element_type=ElementType.TEXT,
                                content=markdown_text,
                                page_number=1,
                            )
                        )
                except Exception as e:
                    logger.warning(f"Fallback markdown export failed: {e}")

            total_images = sum(1 for el in elements if el.element_type == ElementType.IMAGE)
            metadata["total_images"] = total_images

            total_text = sum(
                len(el.content) for el in elements
                if el.element_type == ElementType.TEXT
            )
            if total_text < 100 and page_count > 0:
                warnings.append(
                    "Document has very little extractable text. "
                    "It may be a scanned document requiring OCR."
                )

        except Exception as e:
            logger.error(f"PDF extraction failed: {e}")
            raise

        raw_text = "\n\n".join(
            el.content for el in elements
            if el.element_type == ElementType.TEXT and el.content.strip()
        )

        return ExtractionResult(
            elements=elements,
            metadata=metadata,
            raw_text=raw_text,
            page_count=page_count,
            word_count=len(raw_text.split()) if raw_text else 0,
            extraction_method="docling_pdf",
            warnings=warnings,
        )

    def _extract_docx(
        self,
        file_path: Path,
        extract_images: bool = True,
    ) -> ExtractionResult:
        """Extract content from DOCX with structure preservation."""
        elements: list[ExtractedElement] = []
        warnings: list[str] = []
        metadata: dict[str, Any] = {"format": "docx"}

        try:
            converter = self._get_converter(file_path)
            result = converter.convert(str(file_path))
            doc = result.document

            metadata.update(self._extract_metadata(result))

            # Track seen image data to avoid duplicates
            seen_image_hashes: set[str] = set()
            page_num = 1
            current_section = ""
            order_counter = 0

            for item, level in doc.iterate_items():
                item_order = self._get_item_order(item, order_counter)
                order_counter += 1

                # Handle headings
                if hasattr(item, 'label') and item.label:
                    label = str(item.label).lower()
                    if 'heading' in label or 'title' in label:
                        if hasattr(item, 'text') and item.text:
                            current_section = item.text
                            elements.append(
                                ExtractedElement(
                                    element_type=ElementType.TEXT,
                                    content=f"## {item.text}",
                                    page_number=page_num,
                                    metadata={"type": "heading", "level": level},
                                )
                            )
                            continue

                # Handle tables
                if hasattr(item, 'label') and item.label and 'table' in str(item.label).lower():
                    if hasattr(item, 'text') and item.text:
                        elements.append(
                            ExtractedElement(
                                element_type=ElementType.TEXT,
                                content=item.text,
                                page_number=page_num,
                                metadata={"type": "table", "section": current_section},
                            )
                        )
                        continue

                # Handle images
                if self._is_picture_item(item):
                    if extract_images:
                        img_element = self._extract_picture_item(item, len(elements), doc)
                        if img_element and img_element.image_data:
                            img_hash = img_element.image_data[:200]
                            if img_hash not in seen_image_hashes:
                                seen_image_hashes.add(img_hash)
                                img_element.metadata["section"] = current_section
                                img_element.metadata["position"] = item_order
                                elements.append(img_element)

                # Handle text
                elif hasattr(item, 'text') and item.text:
                    text = item.text.strip()
                    if text:
                        elements.append(
                            ExtractedElement(
                                element_type=ElementType.TEXT,
                                content=text,
                                page_number=page_num,
                                metadata={"section": current_section},
                            )
                        )

            # Also extract from doc.pictures as fallback
            if extract_images and hasattr(doc, 'pictures'):
                for idx, picture in enumerate(doc.pictures):
                    img_element = self._extract_picture(picture, idx, doc)
                    if img_element and img_element.image_data:
                        img_hash = img_element.image_data[:200]
                        if img_hash not in seen_image_hashes:
                            seen_image_hashes.add(img_hash)
                            img_element.metadata["position"] = 1000 + idx
                            elements.append(img_element)

            # Export markdown with images
            try:
                markdown_text = self._export_markdown_with_images(doc)
                if markdown_text:
                    metadata["markdown"] = markdown_text
            except Exception as e:
                logger.warning(f"Markdown export failed: {e}")
                try:
                    markdown_text = doc.export_to_markdown()
                    if markdown_text:
                        metadata["markdown"] = markdown_text
                except Exception:
                    pass

            total_images = sum(1 for el in elements if el.element_type == ElementType.IMAGE)
            metadata["total_images"] = total_images

        except Exception as e:
            logger.error(f"DOCX extraction failed: {e}")
            raise

        raw_text = "\n\n".join(
            el.content for el in elements
            if el.element_type == ElementType.TEXT and el.content.strip()
        )

        return ExtractionResult(
            elements=elements,
            metadata=metadata,
            raw_text=raw_text,
            page_count=1,
            word_count=len(raw_text.split()) if raw_text else 0,
            extraction_method="docling_docx",
            warnings=warnings,
        )

    def _extract_xlsx(self, file_path: Path) -> ExtractionResult:
        """Extract XLSX as structured JSON with schema detection."""
        metadata: dict[str, Any] = {"format": "xlsx", "format_output": "structured_json"}
        warnings: list[str] = []

        try:
            import openpyxl

            workbook = openpyxl.load_workbook(file_path, data_only=True)
            all_tables: list[StructuredTableData] = []
            total_rows = 0
            total_columns = 0

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]

                # Get data from sheet
                rows_data: list[list[Any]] = []
                for row in sheet.iter_rows(values_only=True):
                    if any(cell is not None for cell in row):
                        rows_data.append(list(row))

                if not rows_data:
                    continue

                # First row as headers
                raw_headers = rows_data[0] if rows_data else []
                headers = [
                    self._normalize_header(str(h) if h else f"column_{i}")
                    for i, h in enumerate(raw_headers)
                ]

                # Handle duplicate headers
                seen: dict[str, int] = {}
                unique_headers = []
                for h in headers:
                    if h in seen:
                        seen[h] += 1
                        unique_headers.append(f"{h}_{seen[h]}")
                    else:
                        seen[h] = 0
                        unique_headers.append(h)
                headers = unique_headers

                # Convert data rows to objects
                json_rows: list[dict[str, Any]] = []
                for row_data in rows_data[1:]:
                    row_dict: dict[str, Any] = {}
                    for i, cell in enumerate(row_data):
                        if i < len(headers):
                            field_name = headers[i]
                            row_dict[field_name] = self._parse_cell_value(
                                str(cell) if cell is not None else "",
                                field_name
                            )
                    if any(v is not None for v in row_dict.values()):
                        json_rows.append(row_dict)

                if not json_rows:
                    continue

                # Detect column types
                column_types: dict[str, str] = {}
                for header in headers:
                    values = [row.get(header) for row in json_rows]
                    column_types[header] = self._detect_column_type(values)

                # Build schema
                schema_columns = []
                for i, header in enumerate(headers):
                    if i < len(raw_headers) and raw_headers[i]:
                        original = str(raw_headers[i])
                    else:
                        original = header
                    schema_columns.append({
                        "name": header,
                        "original_name": original,
                        "type": column_types.get(header, "string"),
                        "nullable": any(row.get(header) is None for row in json_rows),
                        "index": i,
                    })

                schema = TableSchema(columns=schema_columns)

                table = StructuredTableData(
                    name=sheet_name,
                    schema=schema,
                    rows=json_rows,
                    row_count=len(json_rows),
                    column_count=len(headers),
                    page_index=len(all_tables),
                    total_pages=len(workbook.sheetnames),
                )
                all_tables.append(table)

                total_rows += len(json_rows)
                total_columns = max(total_columns, len(headers))

            workbook.close()

            # Build summary
            summary = {
                "total_rows": total_rows,
                "total_sheets": len(all_tables),
                "sheet_names": [t.name for t in all_tables],
                "total_columns": total_columns,
            }
            metadata["xlsx_summary"] = summary

            # Generate markdown tables
            markdown_parts = []
            for table in all_tables:
                if table.rows:
                    markdown_parts.append(f"## {table.name}\n")
                    # Header row
                    headers = list(table.rows[0].keys()) if table.rows else []
                    if headers:
                        markdown_parts.append("| " + " | ".join(headers) + " |")
                        markdown_parts.append("| " + " | ".join(["---"] * len(headers)) + " |")
                        # Data rows
                        for row in table.rows:
                            values = [str(row.get(h, "")) for h in headers]
                            markdown_parts.append("| " + " | ".join(values) + " |")
                    markdown_parts.append("")
            if markdown_parts:
                metadata["markdown"] = "\n".join(markdown_parts)

            # Create text element for raw_text
            text_content = f"Excel Workbook: {len(all_tables)} sheets, {total_rows} total rows"
            elements = [
                ExtractedElement(
                    element_type=ElementType.TEXT,
                    content=text_content,
                    page_number=1,
                )
            ]

        except Exception as e:
            logger.error(f"XLSX extraction failed: {e}")
            raise

        return ExtractionResult(
            elements=elements,
            metadata=metadata,
            raw_text=text_content,
            page_count=len(all_tables),
            word_count=0,
            extraction_method="docling_xlsx",
            warnings=warnings,
            structured_data=all_tables if all_tables else None,
        )

    def _extract_pptx(
        self,
        file_path: Path,
        extract_images: bool = True,
    ) -> ExtractionResult:
        """Extract PPTX with slide-by-slide structure."""
        elements: list[ExtractedElement] = []
        warnings: list[str] = []
        metadata: dict[str, Any] = {"format": "pptx"}

        try:
            converter = self._get_converter(file_path)
            result = converter.convert(str(file_path))
            doc = result.document

            metadata.update(self._extract_metadata(result))

            # Track seen image data to avoid duplicates
            seen_image_hashes: set[str] = set()
            slide_num = 1
            current_slide_content: list[str] = []
            order_counter = 0

            for item, level in doc.iterate_items():
                item_page = self._get_item_page(item)
                item_order = self._get_item_order(item, order_counter)
                order_counter += 1

                # New slide
                if item_page != slide_num:
                    if current_slide_content:
                        elements.append(
                            ExtractedElement(
                                element_type=ElementType.TEXT,
                                content="\n".join(current_slide_content),
                                page_number=slide_num,
                                metadata={"type": "slide"},
                            )
                        )
                    current_slide_content = []
                    slide_num = item_page

                # Handle tables
                if hasattr(item, 'label') and item.label and 'table' in str(item.label).lower():
                    if hasattr(item, 'text') and item.text:
                        current_slide_content.append(item.text)
                        continue

                # Handle images
                if self._is_picture_item(item):
                    if extract_images:
                        img_element = self._extract_picture_item(item, len(elements), doc)
                        if img_element and img_element.image_data:
                            img_hash = img_element.image_data[:200]
                            if img_hash not in seen_image_hashes:
                                seen_image_hashes.add(img_hash)
                                img_element.page_number = slide_num
                                img_element.metadata["slide"] = slide_num
                                img_element.metadata["position"] = item_order
                                elements.append(img_element)

                # Handle text
                elif hasattr(item, 'text') and item.text:
                    text = item.text.strip()
                    if text:
                        current_slide_content.append(text)

            # Add last slide
            if current_slide_content:
                elements.append(
                    ExtractedElement(
                        element_type=ElementType.TEXT,
                        content="\n".join(current_slide_content),
                        page_number=slide_num,
                        metadata={"type": "slide"},
                    )
                )

            # Also extract from doc.pictures as fallback
            if extract_images and hasattr(doc, 'pictures'):
                for idx, picture in enumerate(doc.pictures):
                    picture_page = self._get_item_page(picture)
                    img_element = self._extract_picture(picture, idx, doc)
                    if img_element and img_element.image_data:
                        img_hash = img_element.image_data[:200]
                        if img_hash not in seen_image_hashes:
                            seen_image_hashes.add(img_hash)
                            img_element.page_number = picture_page
                            img_element.metadata["slide"] = picture_page
                            img_element.metadata["position"] = 1000 + idx
                            elements.append(img_element)

            metadata["slide_count"] = slide_num
            total_images = sum(1 for el in elements if el.element_type == ElementType.IMAGE)
            metadata["total_images"] = total_images

            # Export markdown with images
            try:
                markdown_text = self._export_markdown_with_images(doc)
                if markdown_text:
                    metadata["markdown"] = markdown_text
            except Exception as e:
                logger.warning(f"Markdown export failed: {e}")
                try:
                    markdown_text = doc.export_to_markdown()
                    if markdown_text:
                        metadata["markdown"] = markdown_text
                except Exception:
                    pass

        except Exception as e:
            logger.error(f"PPTX extraction failed: {e}")
            raise

        raw_text = "\n\n".join(
            el.content for el in elements
            if el.element_type == ElementType.TEXT and el.content.strip()
        )

        return ExtractionResult(
            elements=elements,
            metadata=metadata,
            raw_text=raw_text,
            page_count=metadata.get("slide_count", 1),
            word_count=len(raw_text.split()) if raw_text else 0,
            extraction_method="docling_pptx",
            warnings=warnings,
        )

    def _extract_html(
        self,
        file_path: Path,
        extract_images: bool = True,
    ) -> ExtractionResult:
        """Extract HTML with structure preservation."""
        elements: list[ExtractedElement] = []
        warnings: list[str] = []
        metadata: dict[str, Any] = {"format": "html"}

        try:
            converter = self._get_converter(file_path)
            result = converter.convert(str(file_path))
            doc = result.document

            metadata.update(self._extract_metadata(result))

            # Track seen image data to avoid duplicates
            seen_image_hashes: set[str] = set()
            order_counter = 0

            # Extract content
            for item, level in doc.iterate_items():
                item_order = self._get_item_order(item, order_counter)
                order_counter += 1

                # Handle tables
                if hasattr(item, 'label') and item.label and 'table' in str(item.label).lower():
                    if hasattr(item, 'text') and item.text:
                        elements.append(
                            ExtractedElement(
                                element_type=ElementType.TEXT,
                                content=item.text,
                                page_number=1,
                                metadata={"type": "table", "html_tag": "table"},
                            )
                        )
                        continue

                if self._is_picture_item(item):
                    if extract_images:
                        img_element = self._extract_picture_item(item, len(elements), doc)
                        if img_element and img_element.image_data:
                            img_hash = img_element.image_data[:200]
                            if img_hash not in seen_image_hashes:
                                seen_image_hashes.add(img_hash)
                                img_element.metadata["position"] = item_order
                                elements.append(img_element)

                elif hasattr(item, 'text') and item.text:
                    text = item.text.strip()
                    if text:
                        el_metadata: dict[str, Any] = {}
                        if hasattr(item, 'label') and item.label:
                            el_metadata["html_tag"] = str(item.label)
                        elements.append(
                            ExtractedElement(
                                element_type=ElementType.TEXT,
                                content=text,
                                page_number=1,
                                metadata=el_metadata,
                            )
                        )

            # Also extract from doc.pictures as fallback
            if extract_images and hasattr(doc, 'pictures'):
                for idx, picture in enumerate(doc.pictures):
                    img_element = self._extract_picture(picture, idx, doc)
                    if img_element and img_element.image_data:
                        img_hash = img_element.image_data[:200]
                        if img_hash not in seen_image_hashes:
                            seen_image_hashes.add(img_hash)
                            img_element.metadata["position"] = 1000 + idx
                            elements.append(img_element)

            # Export markdown with images
            try:
                markdown_text = self._export_markdown_with_images(doc)
                if markdown_text:
                    metadata["markdown"] = markdown_text
            except Exception as e:
                logger.warning(f"Markdown export failed: {e}")
                try:
                    markdown_text = doc.export_to_markdown()
                    if markdown_text:
                        metadata["markdown"] = markdown_text
                except Exception:
                    pass

            total_images = sum(1 for el in elements if el.element_type == ElementType.IMAGE)
            metadata["total_images"] = total_images

        except Exception as e:
            logger.error(f"HTML extraction failed: {e}")
            raise

        raw_text = "\n\n".join(
            el.content for el in elements
            if el.element_type == ElementType.TEXT and el.content.strip()
        )

        return ExtractionResult(
            elements=elements,
            metadata=metadata,
            raw_text=raw_text,
            page_count=1,
            word_count=len(raw_text.split()) if raw_text else 0,
            extraction_method="docling_html",
            warnings=warnings,
        )

    def _extract_markdown(self, file_path: Path) -> ExtractionResult:
        """Extract Markdown with code block preservation."""
        elements: list[ExtractedElement] = []
        warnings: list[str] = []
        metadata: dict[str, Any] = {"format": "markdown"}

        try:
            converter = self._get_converter(file_path)
            result = converter.convert(str(file_path))
            doc = result.document

            metadata.update(self._extract_metadata(result))

            # Also read raw file for code blocks
            with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
                raw_content = f.read()

            # Extract code blocks
            code_blocks = re.findall(r'```(\w*)\n(.*?)```', raw_content, re.DOTALL)
            if code_blocks:
                metadata["code_blocks"] = [
                    {"language": lang or "text", "code": code.strip()}
                    for lang, code in code_blocks
                ]

            # Extract content from Docling
            for item, level in doc.iterate_items():
                if hasattr(item, 'text') and item.text:
                    text = item.text.strip()
                    if text:
                        elements.append(
                            ExtractedElement(
                                element_type=ElementType.TEXT,
                                content=text,
                                page_number=1,
                            )
                        )

            # If no elements, use raw content
            if not elements and raw_content:
                elements.append(
                    ExtractedElement(
                        element_type=ElementType.TEXT,
                        content=raw_content,
                        page_number=1,
                    )
                )

            metadata["markdown"] = raw_content

        except Exception as e:
            logger.error(f"Markdown extraction failed: {e}")
            raise

        raw_text = "\n\n".join(
            el.content for el in elements
            if el.element_type == ElementType.TEXT and el.content.strip()
        )

        return ExtractionResult(
            elements=elements,
            metadata=metadata,
            raw_text=raw_text,
            page_count=1,
            word_count=len(raw_text.split()) if raw_text else 0,
            extraction_method="docling_markdown",
            warnings=warnings,
        )

    def _extract_image(self, file_path: Path) -> ExtractionResult:
        """Extract text from image using OCR."""
        elements: list[ExtractedElement] = []
        warnings: list[str] = []
        metadata: dict[str, Any] = {"format": "image"}

        try:
            converter = self._get_converter(file_path)
            result = converter.convert(str(file_path))
            doc = result.document

            metadata.update(self._extract_metadata(result))
            metadata["ocr_enabled"] = self.enable_ocr

            # Extract OCR text
            ocr_text = ""
            for item, level in doc.iterate_items():
                if hasattr(item, 'text') and item.text:
                    text = item.text.strip()
                    if text:
                        ocr_text += text + "\n"
                        elements.append(
                            ExtractedElement(
                                element_type=ElementType.TEXT,
                                content=text,
                                page_number=1,
                                metadata={"source": "ocr"},
                            )
                        )

            # If no text extracted, try markdown export
            if not ocr_text:
                try:
                    markdown_text = doc.export_to_markdown()
                    if markdown_text and markdown_text.strip():
                        ocr_text = markdown_text
                        elements.append(
                            ExtractedElement(
                                element_type=ElementType.TEXT,
                                content=markdown_text,
                                page_number=1,
                                metadata={"source": "ocr_markdown"},
                            )
                        )
                except Exception:
                    pass

            if not ocr_text:
                warnings.append("No text could be extracted from image. OCR may have failed.")

            # Export markdown
            try:
                markdown_text = doc.export_to_markdown()
                if markdown_text and markdown_text.strip():
                    metadata["markdown"] = markdown_text
            except Exception as e:
                logger.warning(f"Markdown export failed for image: {e}")

            # Also include the image itself as base64
            try:
                with open(file_path, 'rb') as f:
                    img_bytes = f.read()
                    b64 = base64.b64encode(img_bytes).decode('utf-8')
                    ext = file_path.suffix.lower()
                    mime = {
                        '.png': 'image/png',
                        '.jpg': 'image/jpeg',
                        '.jpeg': 'image/jpeg',
                        '.gif': 'image/gif',
                        '.webp': 'image/webp',
                        '.bmp': 'image/bmp',
                        '.tiff': 'image/tiff',
                        '.tif': 'image/tiff',
                    }.get(ext, 'image/png')

                    elements.append(
                        ExtractedElement(
                            element_type=ElementType.IMAGE,
                            content=f"Original image: {file_path.name}",
                            page_number=1,
                            image_data=f"data:{mime};base64,{b64}",
                            metadata={"image_type": "original"},
                        )
                    )
            except Exception as e:
                logger.warning(f"Failed to encode image: {e}")

        except Exception as e:
            logger.error(f"Image OCR extraction failed: {e}")
            raise

        raw_text = ocr_text.strip()

        return ExtractionResult(
            elements=elements,
            metadata=metadata,
            raw_text=raw_text,
            page_count=1,
            word_count=len(raw_text.split()) if raw_text else 0,
            extraction_method="docling_image_ocr",
            warnings=warnings,
        )

    def _extract_csv(self, file_path: Path) -> ExtractionResult:
        """Extract CSV as structured JSON with schema detection."""
        metadata: dict[str, Any] = {"format": "csv", "format_output": "structured_json"}
        warnings: list[str] = []

        all_rows: list[dict[str, Any]] = []
        headers: list[str] = []
        original_headers: list[str] = []
        delimiter = ','
        encoding = 'utf-8'

        # Detect encoding
        try:
            with open(file_path, 'rb') as f:
                raw = f.read(8192)
                if raw.startswith(b'\xef\xbb\xbf'):
                    encoding = 'utf-8-sig'
                elif raw.startswith(b'\xff\xfe'):
                    encoding = 'utf-16-le'
                elif raw.startswith(b'\xfe\xff'):
                    encoding = 'utf-16-be'
        except Exception:
            encoding = 'utf-8'

        # Parse CSV
        try:
            with open(file_path, 'r', encoding=encoding, errors='replace') as f:
                sample = f.read(8192)
                f.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample)
                    delimiter = dialect.delimiter
                except csv.Error:
                    delimiter = ','

                reader = csv.reader(f, delimiter=delimiter)

                header_row = next(reader, None)
                if header_row:
                    original_headers = [h.strip() for h in header_row]
                    headers = [self._normalize_header(h) for h in original_headers]

                    seen: dict[str, int] = {}
                    unique_headers = []
                    for h in headers:
                        if h in seen:
                            seen[h] += 1
                            unique_headers.append(f"{h}_{seen[h]}")
                        else:
                            seen[h] = 0
                            unique_headers.append(h)
                    headers = unique_headers

                phone_fields = [h for h in headers if self._is_phone_field(h)]

                for row in reader:
                    if not row or all(not cell.strip() for cell in row):
                        continue

                    row_dict: dict[str, Any] = {}
                    for i, cell in enumerate(row):
                        if i < len(headers):
                            field_name = headers[i]
                            parsed_value = self._parse_cell_value(cell, field_name)
                            row_dict[field_name] = parsed_value

                            if field_name in phone_fields and cell.strip():
                                row_dict[f"{field_name}_clean"] = self._clean_phone(cell)
                        else:
                            row_dict[f"column_{i}"] = self._parse_cell_value(cell, "")

                    all_rows.append(row_dict)

        except Exception as e:
            logger.error(f"Failed to parse CSV: {e}")
            raise

        if not all_rows:
            warnings.append("No data rows found in CSV")
            return ExtractionResult(
                elements=[],
                metadata=metadata,
                raw_text="",
                page_count=0,
                word_count=0,
                extraction_method="docling_csv",
                warnings=warnings,
            )

        # Detect column types
        column_types: dict[str, str] = {}
        column_nullable: dict[str, bool] = {}

        for header in headers:
            values = [row.get(header) for row in all_rows]
            column_types[header] = self._detect_column_type(values)
            column_nullable[header] = any(v is None for v in values)

        # Build schema
        schema_columns = []
        for i, header in enumerate(headers):
            original = original_headers[i] if i < len(original_headers) else header
            col_schema = {
                "name": header,
                "original_name": original,
                "type": column_types.get(header, "string"),
                "nullable": column_nullable.get(header, True),
                "index": i,
            }
            schema_columns.append(col_schema)

        schema = TableSchema(columns=schema_columns)

        # Paginate
        rows_per_page = 500
        total_rows = len(all_rows)
        total_pages = (total_rows + rows_per_page - 1) // rows_per_page

        tables: list[StructuredTableData] = []
        for page_idx in range(total_pages):
            start_idx = page_idx * rows_per_page
            end_idx = min(start_idx + rows_per_page, total_rows)
            page_rows = all_rows[start_idx:end_idx]

            table = StructuredTableData(
                name=file_path.stem,
                schema=schema,
                rows=page_rows,
                row_count=len(page_rows),
                column_count=len(headers),
                page_index=page_idx,
                total_pages=total_pages,
            )
            tables.append(table)

        summary = {
            "total_rows": total_rows,
            "total_columns": len(headers),
            "columns": headers,
            "column_types": column_types,
            "pages": total_pages,
            "rows_per_page": rows_per_page,
            "origin": {
                "source_type": "csv",
                "delimiter": delimiter,
                "encoding": encoding,
            },
        }
        metadata["csv_summary"] = summary

        # Generate markdown table
        markdown_parts = []
        markdown_parts.append(f"## {file_path.stem}\n")
        if headers:
            markdown_parts.append("| " + " | ".join(headers) + " |")
            markdown_parts.append("| " + " | ".join(["---"] * len(headers)) + " |")
            # Limit to first 100 rows for markdown preview
            for row in all_rows[:100]:
                values = [str(row.get(h, "")) for h in headers]
                markdown_parts.append("| " + " | ".join(values) + " |")
            if total_rows > 100:
                markdown_parts.append(f"\n*... and {total_rows - 100} more rows*")
        metadata["markdown"] = "\n".join(markdown_parts)

        text_content = (
            f"CSV Data: {total_rows} rows, {len(headers)} columns. "
            f"Columns: {', '.join(headers)}"
        )
        elements = [
            ExtractedElement(
                element_type=ElementType.TEXT,
                content=text_content,
                page_number=1,
            )
        ]

        return ExtractionResult(
            elements=elements,
            metadata=metadata,
            raw_text=text_content,
            page_count=total_pages,
            word_count=0,
            extraction_method="docling_csv",
            warnings=warnings,
            structured_data=tables,
        )

    def _extract_asciidoc(self, file_path: Path) -> ExtractionResult:
        """Extract AsciiDoc content."""
        elements: list[ExtractedElement] = []
        warnings: list[str] = []
        metadata: dict[str, Any] = {"format": "asciidoc"}

        try:
            converter = self._get_converter(file_path)
            result = converter.convert(str(file_path))
            doc = result.document

            metadata.update(self._extract_metadata(result))

            for item, level in doc.iterate_items():
                if hasattr(item, 'text') and item.text:
                    text = item.text.strip()
                    if text:
                        elements.append(
                            ExtractedElement(
                                element_type=ElementType.TEXT,
                                content=text,
                                page_number=1,
                            )
                        )

            # Fallback to raw file
            if not elements:
                with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
                    raw_content = f.read()
                if raw_content:
                    elements.append(
                        ExtractedElement(
                            element_type=ElementType.TEXT,
                            content=raw_content,
                            page_number=1,
                        )
                    )

            # Export markdown
            try:
                markdown_text = doc.export_to_markdown()
                if markdown_text and markdown_text.strip():
                    metadata["markdown"] = markdown_text
            except Exception as e:
                logger.warning(f"Markdown export failed for AsciiDoc: {e}")

        except Exception as e:
            logger.error(f"AsciiDoc extraction failed: {e}")
            raise

        raw_text = "\n\n".join(
            el.content for el in elements
            if el.element_type == ElementType.TEXT and el.content.strip()
        )

        return ExtractionResult(
            elements=elements,
            metadata=metadata,
            raw_text=raw_text,
            page_count=1,
            word_count=len(raw_text.split()) if raw_text else 0,
            extraction_method="docling_asciidoc",
            warnings=warnings,
        )

    def _extract_generic(
        self,
        file_path: Path,
        extract_images: bool = True,
    ) -> ExtractionResult:
        """Generic extraction for any supported format."""
        elements: list[ExtractedElement] = []
        warnings: list[str] = []
        metadata: dict[str, Any] = {"format": file_path.suffix.lower()[1:]}

        try:
            converter = self._get_converter(file_path)
            result = converter.convert(str(file_path))
            doc = result.document

            metadata.update(self._extract_metadata(result))

            for item, level in doc.iterate_items():
                if self._is_picture_item(item):
                    if extract_images:
                        img_element = self._extract_picture_item(item, len(elements), doc)
                        if img_element:
                            elements.append(img_element)

                elif hasattr(item, 'text') and item.text:
                    text = item.text.strip()
                    if text:
                        elements.append(
                            ExtractedElement(
                                element_type=ElementType.TEXT,
                                content=text,
                                page_number=1,
                            )
                        )

            try:
                markdown_text = doc.export_to_markdown()
                if markdown_text:
                    metadata["markdown"] = markdown_text
            except Exception:
                pass

        except Exception as e:
            logger.error(f"Generic extraction failed: {e}")
            raise

        raw_text = "\n\n".join(
            el.content for el in elements
            if el.element_type == ElementType.TEXT and el.content.strip()
        )

        return ExtractionResult(
            elements=elements,
            metadata=metadata,
            raw_text=raw_text,
            page_count=1,
            word_count=len(raw_text.split()) if raw_text else 0,
            extraction_method=f"docling_{file_path.suffix.lower()[1:]}",
            warnings=warnings,
        )

    # Helper methods

    def _is_picture_item(self, item) -> bool:
        """Check if an item is a picture/image element."""
        item_type = type(item).__name__
        if item_type in ('PictureItem', 'Picture'):
            return True
        return hasattr(item, 'image') and item.image

    def _extract_metadata(self, result) -> dict[str, Any]:
        """Extract document metadata from Docling result."""
        metadata: dict[str, Any] = {}
        try:
            if result.input:
                if result.input.page_count:
                    metadata["page_count"] = result.input.page_count
                if result.input.filesize:
                    metadata["file_size"] = result.input.filesize
                if result.input.format:
                    metadata["docling_format"] = str(result.input.format)
        except Exception as e:
            logger.warning(f"Failed to extract metadata: {e}")
        return metadata

    def _export_markdown_with_images(self, doc) -> str:
        """Export document to markdown with embedded base64 images."""
        try:
            from docling_core.transforms.serializer.markdown import (
                MarkdownDocSerializer,
                MarkdownParams,
            )
            params = MarkdownParams(image_mode=_ImageRefMode.EMBEDDED)
            serializer = MarkdownDocSerializer(doc=doc, params=params)
            result = serializer.serialize()
            return result.text
        except ImportError:
            return doc.export_to_markdown()
        except Exception as e:
            logger.warning(f"Markdown serializer failed: {e}")
            return doc.export_to_markdown()

    def _get_item_page(self, item) -> int:
        """Get page number for a document item."""
        try:
            if hasattr(item, 'prov') and item.prov and len(item.prov) > 0:
                prov = item.prov[0]
                if hasattr(prov, 'page_no'):
                    return prov.page_no
        except Exception:
            pass
        return 1

    def _get_item_order(self, item, default_order: int) -> int:
        """Get the order/position of an item within its page.

        In PDF coordinate systems, Y=0 is typically at the bottom and increases upward.
        So elements at the top of the page have HIGH Y/top values.
        We negate the value so that sorting ascending gives us top-to-bottom order.
        """
        try:
            if hasattr(item, 'prov') and item.prov and len(item.prov) > 0:
                prov = item.prov[0]
                if hasattr(prov, 'bbox') and prov.bbox:
                    bbox = prov.bbox
                    # Negate to get top-to-bottom order when sorting ascending
                    if hasattr(bbox, 't'):
                        return -int(bbox.t * 1000)
                    elif hasattr(bbox, 'top'):
                        return -int(bbox.top * 1000)
                    elif hasattr(bbox, 'y'):
                        return -int(bbox.y * 1000)
        except Exception:
            pass
        return default_order

    def _extract_picture_item(self, item, index: int, doc) -> Optional[ExtractedElement]:
        """Extract a picture from a document item."""
        try:
            page_num = self._get_item_page(item)

            image = None
            if hasattr(item, 'image') and item.image:
                image = item.image
            elif hasattr(item, 'get_image'):
                try:
                    image = item.get_image(doc)
                except Exception:
                    pass

            if not image:
                return None

            image_data = None
            if hasattr(image, 'uri') and image.uri:
                uri = str(image.uri)
                if uri.startswith('data:'):
                    image_data = uri
                else:
                    try:
                        img_path = Path(uri)
                        if img_path.exists():
                            with open(img_path, 'rb') as f:
                                img_bytes = f.read()
                                b64 = base64.b64encode(img_bytes).decode('utf-8')
                                ext = img_path.suffix.lower()
                                mime = {
                                    '.png': 'image/png',
                                    '.jpg': 'image/jpeg',
                                    '.jpeg': 'image/jpeg',
                                    '.gif': 'image/gif',
                                    '.webp': 'image/webp',
                                }.get(ext, 'image/png')
                                image_data = f"data:{mime};base64,{b64}"
                    except Exception:
                        pass

            if not image_data:
                return None

            caption = ""
            if hasattr(item, 'caption_text'):
                try:
                    caption = item.caption_text(doc=doc) or ""
                except Exception:
                    pass
            elif hasattr(item, 'text') and item.text:
                caption = item.text

            width = getattr(image, 'width', None)
            height = getattr(image, 'height', None)

            return ExtractedElement(
                element_type=ElementType.IMAGE,
                content=caption or f"Image {index + 1} on page {page_num}",
                page_number=page_num,
                image_data=image_data,
                metadata={
                    "image_index": index,
                    "width": width,
                    "height": height,
                    "image_type": "docling_picture",
                },
            )

        except Exception as e:
            logger.warning(f"Failed to extract picture item {index}: {e}")
        return None

    def _extract_picture(self, picture, index: int, doc) -> Optional[ExtractedElement]:
        """Extract a picture from the document."""
        try:
            page_num = self._get_item_page(picture)

            if hasattr(picture, 'image') and picture.image:
                image = picture.image

                if hasattr(image, 'uri') and image.uri:
                    uri = str(image.uri)

                    if uri.startswith('data:'):
                        image_data = uri
                    else:
                        try:
                            img_path = Path(uri)
                            if img_path.exists():
                                with open(img_path, 'rb') as f:
                                    img_bytes = f.read()
                                    b64 = base64.b64encode(img_bytes).decode('utf-8')
                                    ext = img_path.suffix.lower()
                                    mime = {
                                        '.png': 'image/png',
                                        '.jpg': 'image/jpeg',
                                        '.jpeg': 'image/jpeg',
                                        '.gif': 'image/gif',
                                        '.webp': 'image/webp',
                                    }.get(ext, 'image/png')
                                    image_data = f"data:{mime};base64,{b64}"
                            else:
                                return None
                        except Exception:
                            return None

                    caption = ""
                    if hasattr(picture, 'caption_text'):
                        try:
                            caption = picture.caption_text(doc=doc) or ""
                        except Exception:
                            pass

                    width = getattr(image, 'width', None)
                    height = getattr(image, 'height', None)

                    return ExtractedElement(
                        element_type=ElementType.IMAGE,
                        content=caption or f"Image {index + 1} on page {page_num}",
                        page_number=page_num,
                        image_data=image_data,
                        metadata={
                            "image_index": index,
                            "width": width,
                            "height": height,
                            "image_type": "docling_picture",
                        },
                    )

        except Exception as e:
            logger.warning(f"Failed to extract picture {index}: {e}")
        return None

    def _normalize_header(self, header: str) -> str:
        """Normalize header to snake_case."""
        if not header:
            return "column"
        clean = header.strip().lower()
        clean = re.sub(r'[^\w\s]', '', clean)
        clean = re.sub(r'\s+', '_', clean)
        clean = re.sub(r'_+', '_', clean)
        clean = clean.strip('_')
        return clean or "column"

    def _detect_column_type(self, values: list[Any]) -> str:
        """Detect the data type for a column based on sample values."""
        non_empty = [v for v in values if v is not None and str(v).strip()]
        if not non_empty:
            return "string"

        sample = non_empty[:100]
        int_count = 0
        float_count = 0
        bool_count = 0
        date_count = 0

        date_patterns = [
            r'^\d{4}-\d{2}-\d{2}$',
            r'^\d{2}/\d{2}/\d{4}$',
            r'^\d{2}-\d{2}-\d{4}$',
        ]

        for val in sample:
            s = str(val).strip()

            if s.lower() in ('true', 'false', 'yes', 'no', '1', '0'):
                bool_count += 1
                continue

            for pattern in date_patterns:
                if re.match(pattern, s):
                    date_count += 1
                    break
            else:
                try:
                    if s.isdigit() or (s.startswith('-') and s[1:].isdigit()):
                        int_count += 1
                        continue
                except (ValueError, IndexError):
                    pass

                try:
                    float(s)
                    if '.' in s or 'e' in s.lower():
                        float_count += 1
                        continue
                except ValueError:
                    pass

        total = len(sample)
        threshold = 0.8

        if date_count / total >= threshold:
            return "date"
        if int_count / total >= threshold:
            return "integer"
        if float_count / total >= threshold:
            return "number"
        if bool_count / total >= threshold:
            return "boolean"

        return "string"

    def _is_id_or_phone_field(self, field_name: str) -> bool:
        """Check if field should always be treated as string."""
        id_patterns = [
            'id', 'phone', 'tel', 'fax', 'mobile',
            'zip', 'postal', 'code', 'sku', 'isbn'
        ]
        field_lower = field_name.lower()
        return any(pattern in field_lower for pattern in id_patterns)

    def _is_phone_field(self, field_name: str) -> bool:
        """Check if field is specifically a phone number field."""
        phone_patterns = ['phone', 'tel', 'fax', 'mobile', 'cell']
        field_lower = field_name.lower()
        return any(pattern in field_lower for pattern in phone_patterns)

    def _clean_phone(self, value: str) -> str:
        """Extract digits only from phone number."""
        if not value:
            return ""
        return re.sub(r'[^\d]', '', value)

    def _parse_cell_value(
        self, value: str, field_name: str = "", force_string: bool = False
    ) -> Any:
        """Parse cell value to appropriate type."""
        if value is None:
            return None

        if not isinstance(value, str):
            value = str(value)

        value = value.strip()
        if not value:
            return None

        if force_string or self._is_id_or_phone_field(field_name):
            return value

        if value.lower() in ('true', 'yes'):
            return True
        if value.lower() in ('false', 'no'):
            return False

        if not value.startswith('0') or value == '0':
            try:
                if value.isdigit() or (value.startswith('-') and value[1:].isdigit()):
                    return int(value)
            except (ValueError, IndexError):
                pass

        try:
            if '.' in value and not value.startswith('0.') or 'e' in value.lower():
                parsed = float(value)
                if '.' in value:
                    parts = value.split('.')
                    if len(parts) == 2 and parts[1].isdigit():
                        return parsed
        except ValueError:
            pass

        return value


# Backward compatibility alias
DoclingPDFExtractor = DoclingExtractor
